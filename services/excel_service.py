# [V9 - Excel] ExcelService — génération et lecture de fichiers Excel pour Al Qalam.
#
# Utilise la bibliothèque `openpyxl` :
#   - Workbook / Worksheet    : création et manipulation de classeurs
#   - PatternFill             : colorisation conditionnelle des cellules
#   - Font / Alignment / Border : mise en forme professionnelle
#   - BarChart / Reference    : graphique en barres par catégorie
#   - load_workbook           : lecture d'un bon de commande fournisseur
#
# FONCTIONNALITÉS V9 :
#   1. exporter_rapport_stock(chemin) → Workbook multi-feuilles
#      Feuille 1 "Catalogue"     : tableau colorisé (vert/orange/rouge selon stock)
#      Feuille 2 "Ruptures"      : produits en rupture uniquement
#      Feuille 3 "Statistiques"  : totaux par catégorie + graphique en barres
#
#   2. importer_bon_commande(chemin) → dict rapport
#      Lit un bon de commande Excel : colonnes ref, qte (+ note optionnelle)
#      Pour chaque ligne valide : entree_stock() dans le StockService

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib  import Path

from config import (
    EXCEL_DIR,
    EXCEL_COULEUR_ENTETE, EXCEL_COULEUR_RUPTURE,
    EXCEL_COULEUR_ALERTE, EXCEL_COULEUR_OK, EXCEL_COULEUR_TOTAL,
)
from decorateurs.journalisation import journaliser


# ── Styles réutilisables ───────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    """Crée un PatternFill de fond uni à partir d'un code hexadécimal."""
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font_entete() -> Font:
    """Police blanche, grasse, taille 11 — pour les en-têtes."""
    return Font(name="Calibri", bold=True, color="FFFFFF", size=11)

def _font_titre() -> Font:
    """Police titre feuille : bleu foncé, grasse, taille 14."""
    return Font(name="Calibri", bold=True, color="1F4E79", size=14)

def _font_total() -> Font:
    """Police ligne total : grasse, taille 11."""
    return Font(name="Calibri", bold=True, size=11)

def _aligner(horizontal="center", vertical="center") -> Alignment:
    """Alignement standard centré (wrap_text=True)."""
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)

def _bordure_fine() -> Border:
    """Bordure fine sur les quatre côtés."""
    fin = Side(style="thin", color="BDBDBD")
    return Border(left=fin, right=fin, top=fin, bottom=fin)

def _appliquer_style_entete(cellule, couleur_fond=None):
    """Applique le style complet d'un en-tête au classeur."""
    if couleur_fond is None:
        couleur_fond = EXCEL_COULEUR_ENTETE
    cellule.fill      = _fill(couleur_fond)
    cellule.font      = _font_entete()
    cellule.alignment = _aligner()
    cellule.border    = _bordure_fine()

def _appliquer_style_donnee(cellule, couleur_fond: str):
    """Applique fond coloré + alignement + bordure à une cellule de données."""
    cellule.fill      = _fill(couleur_fond)
    cellule.alignment = _aligner(horizontal="left")
    cellule.border    = _bordure_fine()


class ErreurExcel(Exception):
    """Levée quand un fichier Excel ne respecte pas la structure attendue."""


class ExcelService:
    """
    Service d'import/export Excel pour Al Qalam.

    [V9] Utilise openpyxl pour :
      - Générer un rapport de stock multi-feuilles colorisé
      - Lire un bon de commande fournisseur au format Excel

    Usage :
        service = ExcelService(stock_service)
        service.exporter_rapport_stock()         # rapport complet
        rapport = service.importer_bon_commande("bc_fournisseur.xlsx")
    """

    # Colonnes de la feuille Catalogue
    COLONNES_CATALOGUE = [
        ("Référence",   12),
        ("Nom",         28),
        ("Catégorie",   16),
        ("Qté",          8),
        ("Seuil min",   10),
        ("P. Achat",    12),
        ("P. Vente",    12),
        ("Valeur stock",14),
        ("Marge unit.", 12),
        ("Statut",      12),
    ]

    # Colonnes du bon de commande attendu à l'import
    COLONNES_BON_CMD = ["ref", "qte"]

    def __init__(self, stock_service):
        self._stock = stock_service
        EXCEL_DIR.mkdir(parents=True, exist_ok=True)

    # ══════════════════════════════════════════════════════════════════════════
    # EXPORT — Rapport stock multi-feuilles
    # ══════════════════════════════════════════════════════════════════════════

    @journaliser("export_rapport_excel")
    def exporter_rapport_stock(self, chemin: str | Path | None = None) -> Path:
        """
        Génère un classeur Excel multi-feuilles avec rapport complet du stock.

        Feuilles créées :
          - "📦 Catalogue"    : tous les produits, colorisés selon l'état du stock
          - "⚠️ Ruptures"     : produits avec qte = 0 ou qte ≤ seuil_min
          - "📊 Statistiques" : résumé par catégorie + graphique en barres

        Colorisation conditionnelle (Feuille Catalogue) :
          - Rouge pâle  : stock = 0 (rupture totale)
          - Orange pâle : 0 < stock ≤ seuil_min (alerte)
          - Vert pâle   : stock > seuil_min (normal)

        Args:
            chemin : chemin du fichier .xlsx (défaut : data/excel/rapport_YYYYMMDD.xlsx)

        Returns:
            Path du fichier créé
        """
        if chemin is None:
            horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
            chemin = EXCEL_DIR / f"rapport_stock_{horodatage}.xlsx"
        chemin = Path(chemin)
        chemin.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # supprimer la feuille vide par défaut

        produits = sorted(self._stock.lister_tous())

        # ── Feuille 1 : Catalogue ──────────────────────────────────────────
        ws_cat = wb.create_sheet("📦 Catalogue")
        self._remplir_catalogue(ws_cat, produits)

        # ── Feuille 2 : Ruptures ───────────────────────────────────────────
        ws_rup = wb.create_sheet("⚠️ Ruptures")
        ruptures = [p for p in produits if p.qte <= p.seuil_min]
        self._remplir_ruptures(ws_rup, ruptures)

        # ── Feuille 3 : Statistiques ───────────────────────────────────────
        ws_stat = wb.create_sheet("📊 Statistiques")
        self._remplir_statistiques(ws_stat, produits)

        wb.save(chemin)
        return chemin

    # ── Feuille 1 : Catalogue ─────────────────────────────────────────────

    def _remplir_catalogue(self, ws, produits: list):
        """Remplit la feuille Catalogue avec colorisation conditionnelle."""
        # Titre
        ws.merge_cells("A1:J1")
        titre = ws["A1"]
        titre.value     = f"📦  Al Qalam — Catalogue de stock  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        titre.font      = _font_titre()
        titre.alignment = _aligner(horizontal="center")
        ws.row_dimensions[1].height = 28

        # En-têtes (ligne 2)
        for col_idx, (label, largeur) in enumerate(self.COLONNES_CATALOGUE, start=1):
            cellule = ws.cell(row=2, column=col_idx, value=label)
            _appliquer_style_entete(cellule)
            ws.column_dimensions[get_column_letter(col_idx)].width = largeur
        ws.row_dimensions[2].height = 22

        # Données (à partir de la ligne 3)
        valeur_totale = 0.0
        for idx, p in enumerate(produits):
            ligne = idx + 3

            # Déterminer la couleur selon l'état du stock
            if p.qte == 0:
                couleur = EXCEL_COULEUR_RUPTURE
                statut  = "❌ Rupture"
            elif p.qte <= p.seuil_min:
                couleur = EXCEL_COULEUR_ALERTE
                statut  = "⚠️ Alerte"
            else:
                couleur = EXCEL_COULEUR_OK
                statut  = "✅ Normal"

            valeur_stock = p.valeur_stock()
            valeur_totale += valeur_stock

            valeurs = [
                p.ref,
                p.nom,
                p.categorie,
                p.qte,
                p.seuil_min,
                round(p.prix_achat, 3),
                round(p.prix_vente, 3),
                round(valeur_stock, 3),
                round(p.marge_unitaire(), 3),
                statut,
            ]
            for col_idx, valeur in enumerate(valeurs, start=1):
                cellule = ws.cell(row=ligne, column=col_idx, value=valeur)
                _appliquer_style_donnee(cellule, couleur)
                # Centrer les colonnes numériques
                if col_idx in (4, 5, 6, 7, 8, 9):
                    cellule.alignment = _aligner(horizontal="center")

            ws.row_dimensions[ligne].height = 18

        # Ligne totaux
        ligne_total = len(produits) + 3
        ws.merge_cells(f"A{ligne_total}:G{ligne_total}")
        cell_lbl = ws[f"A{ligne_total}"]
        cell_lbl.value     = f"TOTAL  ({len(produits)} produits)"
        cell_lbl.font      = _font_total()
        cell_lbl.fill      = _fill(EXCEL_COULEUR_TOTAL)
        cell_lbl.alignment = _aligner(horizontal="right")
        cell_lbl.border    = _bordure_fine()

        cell_val = ws.cell(row=ligne_total, column=8, value=round(valeur_totale, 3))
        cell_val.font      = _font_total()
        cell_val.fill      = _fill(EXCEL_COULEUR_TOTAL)
        cell_val.alignment = _aligner(horizontal="center")
        cell_val.border    = _bordure_fine()

        # Figer la ligne d'en-tête
        ws.freeze_panes = "A3"

    # ── Feuille 2 : Ruptures ──────────────────────────────────────────────

    def _remplir_ruptures(self, ws, ruptures: list):
        """Remplit la feuille Ruptures avec les produits en alerte ou rupture."""
        ws.merge_cells("A1:F1")
        titre = ws["A1"]
        titre.value     = f"⚠️  Produits en rupture / alerte  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        titre.font      = _font_titre()
        titre.alignment = _aligner(horizontal="center")
        ws.row_dimensions[1].height = 28

        entetes = [("Référence", 12), ("Nom", 28), ("Catégorie", 16),
                   ("Qté", 8), ("Seuil min", 10), ("Statut", 14)]
        for col_idx, (label, largeur) in enumerate(entetes, start=1):
            cellule = ws.cell(row=2, column=col_idx, value=label)
            _appliquer_style_entete(cellule, couleur_fond="C0392B")  # rouge foncé
            ws.column_dimensions[get_column_letter(col_idx)].width = largeur
        ws.row_dimensions[2].height = 22

        if not ruptures:
            ws.merge_cells("A3:F3")
            cell = ws["A3"]
            cell.value     = "✅ Aucune rupture ni alerte — stock en parfait état !"
            cell.font      = Font(name="Calibri", italic=True, color="27AE60", size=11)
            cell.alignment = _aligner(horizontal="center")
            return

        for idx, p in enumerate(ruptures):
            ligne  = idx + 3
            couleur = EXCEL_COULEUR_RUPTURE if p.qte == 0 else EXCEL_COULEUR_ALERTE
            statut  = "❌ Rupture" if p.qte == 0 else "⚠️ Alerte"
            valeurs = [p.ref, p.nom, p.categorie, p.qte, p.seuil_min, statut]
            for col_idx, valeur in enumerate(valeurs, start=1):
                cellule = ws.cell(row=ligne, column=col_idx, value=valeur)
                _appliquer_style_donnee(cellule, couleur)
                if col_idx in (4, 5):
                    cellule.alignment = _aligner(horizontal="center")
            ws.row_dimensions[ligne].height = 18

        ws.freeze_panes = "A3"

    # ── Feuille 3 : Statistiques ──────────────────────────────────────────

    def _remplir_statistiques(self, ws, produits: list):
        """Remplit la feuille Statistiques avec les totaux par catégorie + graphique."""
        ws.merge_cells("A1:F1")
        titre = ws["A1"]
        titre.value     = f"📊  Statistiques par catégorie  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        titre.font      = _font_titre()
        titre.alignment = _aligner(horizontal="center")
        ws.row_dimensions[1].height = 28

        entetes = [
            ("Catégorie", 18), ("Nb produits", 12), ("Valeur stock", 14),
            ("Marge totale", 14), ("Ruptures", 10), ("Alertes", 10),
        ]
        for col_idx, (label, largeur) in enumerate(entetes, start=1):
            cellule = ws.cell(row=2, column=col_idx, value=label)
            _appliquer_style_entete(cellule)
            ws.column_dimensions[get_column_letter(col_idx)].width = largeur
        ws.row_dimensions[2].height = 22

        # Agrégation par catégorie
        stats: dict[str, dict] = {}
        for p in produits:
            cat = p.categorie
            if cat not in stats:
                stats[cat] = {
                    "nb": 0, "valeur": 0.0, "marge": 0.0,
                    "ruptures": 0, "alertes": 0,
                }
            stats[cat]["nb"]     += 1
            stats[cat]["valeur"] += p.valeur_stock()
            stats[cat]["marge"]  += p.qte * p.marge_unitaire()
            if p.qte == 0:
                stats[cat]["ruptures"] += 1
            elif p.qte <= p.seuil_min:
                stats[cat]["alertes"] += 1

        ligne_debut_donnees = 3
        for idx, (cat, s) in enumerate(sorted(stats.items())):
            ligne = idx + ligne_debut_donnees
            couleur = EXCEL_COULEUR_OK
            if s["ruptures"] > 0:
                couleur = EXCEL_COULEUR_RUPTURE
            elif s["alertes"] > 0:
                couleur = EXCEL_COULEUR_ALERTE
            valeurs = [
                cat, s["nb"],
                round(s["valeur"], 3), round(s["marge"], 3),
                s["ruptures"], s["alertes"],
            ]
            for col_idx, valeur in enumerate(valeurs, start=1):
                cellule = ws.cell(row=ligne, column=col_idx, value=valeur)
                _appliquer_style_donnee(cellule, couleur)
                if col_idx > 1:
                    cellule.alignment = _aligner(horizontal="center")
            ws.row_dimensions[ligne].height = 18

        # Ligne total général
        nb_cats   = len(stats)
        ligne_tot = ligne_debut_donnees + nb_cats
        ws.merge_cells(f"A{ligne_tot}:A{ligne_tot}")
        cell_lbl = ws.cell(row=ligne_tot, column=1, value="TOTAL GÉNÉRAL")
        cell_lbl.font      = _font_total()
        cell_lbl.fill      = _fill(EXCEL_COULEUR_TOTAL)
        cell_lbl.alignment = _aligner(horizontal="center")
        cell_lbl.border    = _bordure_fine()

        totaux = [
            sum(s["nb"]       for s in stats.values()),
            round(sum(s["valeur"]  for s in stats.values()), 3),
            round(sum(s["marge"]   for s in stats.values()), 3),
            sum(s["ruptures"] for s in stats.values()),
            sum(s["alertes"]  for s in stats.values()),
        ]
        for col_idx, valeur in enumerate(totaux, start=2):
            cellule = ws.cell(row=ligne_tot, column=col_idx, value=valeur)
            cellule.font      = _font_total()
            cellule.fill      = _fill(EXCEL_COULEUR_TOTAL)
            cellule.alignment = _aligner(horizontal="center")
            cellule.border    = _bordure_fine()

        # ── Graphique en barres : Valeur de stock par catégorie ───────────
        if nb_cats >= 1:
            chart = BarChart()
            chart.type    = "col"       # barres verticales
            chart.title   = "Valeur de stock par catégorie"
            chart.y_axis.title = "Valeur (TND)"
            chart.x_axis.title = "Catégorie"
            chart.style   = 10
            chart.width   = 20
            chart.height  = 12

            # Données : colonne C (valeur stock), lignes 3 → 3+nb_cats-1
            data = Reference(
                ws,
                min_col=3, max_col=3,
                min_row=2, max_row=2 + nb_cats,   # inclut l'en-tête
            )
            cats = Reference(
                ws,
                min_col=1,
                min_row=ligne_debut_donnees,
                max_row=ligne_debut_donnees + nb_cats - 1,
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, f"A{ligne_tot + 2}")

        ws.freeze_panes = "A3"

    # ══════════════════════════════════════════════════════════════════════════
    # IMPORT — Bon de commande Excel
    # ══════════════════════════════════════════════════════════════════════════

    @journaliser("import_bon_commande_excel")
    def importer_bon_commande(self, chemin: str | Path) -> dict:
        """
        Importe un bon de commande fournisseur depuis un fichier Excel.

        Structure attendue (première feuille) :
          - Ligne 1    : en-têtes (doit contenir au minimum : ref, qte)
          - Ligne 2+   : données  (une ligne par produit à réceptionner)
          - Colonne "note" : optionnelle (note du mouvement d'entrée)

        Comportement :
          - Pour chaque ligne valide : appelle stock_service.entree_stock(ref, qte, note)
          - ref inconnue du stock    → ajoutée à erreurs[], traitement continue
          - qte non entière/négative → ajoutée à erreurs[], traitement continue

        Args:
            chemin : chemin du fichier .xlsx (str ou Path)

        Returns:
            dict {
                "traites"     : nb de lignes traitées avec succès,
                "erreurs"     : liste de str décrivant les lignes rejetées,
                "total_lues"  : nb total de lignes de données lues,
                "total_qte"   : quantité totale réceptionnée,
            }
        """
        chemin = Path(chemin)
        if not chemin.exists():
            raise FileNotFoundError(f"Fichier introuvable : {chemin}")

        try:
            wb = openpyxl.load_workbook(chemin, data_only=True)
        except Exception as e:
            raise ErreurExcel(f"Impossible de lire le fichier Excel : {e}")

        ws = wb.active

        # ── Lecture des en-têtes (ligne 1) ────────────────────────────────
        entetes = []
        for cell in ws[1]:
            valeur = cell.value
            entetes.append(str(valeur).strip().lower() if valeur is not None else "")

        # Vérification des colonnes requises
        manquantes = [col for col in self.COLONNES_BON_CMD if col not in entetes]
        if manquantes:
            raise ErreurExcel(
                f"Colonnes manquantes dans {chemin.name} : "
                f"{', '.join(manquantes)}\n"
                f"Colonnes requises : {', '.join(self.COLONNES_BON_CMD)}"
            )

        idx_ref  = entetes.index("ref")
        idx_qte  = entetes.index("qte")
        idx_note = entetes.index("note") if "note" in entetes else None

        rapport = {
            "traites": 0, "erreurs": [],
            "total_lues": 0, "total_qte": 0,
        }

        # ── Traitement ligne par ligne (à partir de la ligne 2) ───────────
        for numero_ligne, ligne in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Ignorer les lignes complètement vides
            if all(v is None for v in ligne):
                continue

            rapport["total_lues"] += 1
            try:
                ref  = str(ligne[idx_ref]).strip()  if ligne[idx_ref]  is not None else ""
                qte_raw = ligne[idx_qte]
                note = str(ligne[idx_note]).strip() if idx_note is not None and ligne[idx_note] is not None else "Bon de commande"

                if not ref:
                    raise ValueError("La référence ne peut pas être vide")

                try:
                    qte = int(float(str(qte_raw)))
                except (ValueError, TypeError):
                    raise ValueError(f"Quantité invalide : '{qte_raw}'")

                if qte <= 0:
                    raise ValueError(f"La quantité doit être > 0 (reçu : {qte})")

                # Vérifier que la référence existe dans le stock
                if ref not in self._stock:
                    raise ValueError(f"Référence '{ref}' inconnue dans le stock")

                # Enregistrer l'entrée de stock
                self._stock.entree_stock(ref, qte, note)
                rapport["traites"]   += 1
                rapport["total_qte"] += qte

            except (ValueError, KeyError) as e:
                rapport["erreurs"].append(f"Ligne {numero_ligne} rejetée : {e}")
                continue

        return rapport

    # ── Prévisualisation ──────────────────────────────────────────────────

    def previsualiser_excel(self, chemin: str | Path, nb_lignes: int = 5) -> dict:
        """
        Lit les premières lignes d'un Excel pour prévisualisation avant import.

        Args:
            chemin    : chemin du fichier .xlsx
            nb_lignes : nombre de lignes de données à lire (sans l'en-tête)

        Returns:
            dict {
                "colonnes"  : liste des noms de colonnes (ligne 1),
                "lignes"    : liste de listes (données brutes),
                "total_col" : nombre de colonnes,
                "valide"    : bool (colonnes ref et qte présentes ?),
                "nb_feuilles": nombre de feuilles dans le classeur,
                "feuilles"  : liste des noms de feuilles,
            }
        """
        chemin  = Path(chemin)
        resultat = {
            "colonnes": [], "lignes": [], "total_col": 0,
            "valide": False, "nb_feuilles": 0, "feuilles": [],
        }
        if not chemin.exists():
            return resultat

        try:
            wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
        except Exception:
            return resultat

        resultat["nb_feuilles"] = len(wb.sheetnames)
        resultat["feuilles"]    = wb.sheetnames
        ws = wb.active

        lignes = list(ws.iter_rows(values_only=True))
        if not lignes:
            return resultat

        # Ligne 1 = en-têtes
        colonnes = [str(v).strip() if v is not None else "" for v in lignes[0]]
        resultat["colonnes"]  = colonnes
        resultat["total_col"] = len(colonnes)

        requises = set(self.COLONNES_BON_CMD)
        resultat["valide"] = requises.issubset({c.lower() for c in colonnes})

        # Lignes de données
        for ligne in lignes[1:nb_lignes + 1]:
            resultat["lignes"].append([str(v) if v is not None else "" for v in ligne])

        wb.close()
        return resultat
